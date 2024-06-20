# UtilForKatya
import openpyxl
from openpyxl.styles import NamedStyle, Font
from datetime import timedelta

def format_duration(duration):
    hours, remainder = divmod(duration.total_seconds(), 3600)
    minutes, _ = divmod(remainder, 60)
    return f"{int(hours):02}:{int(minutes):02}"

def is_of_type(event_type, event_types):
    return any(event_type.startswith(t) for t in event_types)

def main():
    file_path = r"C:\Users\ekserova\report.xlsx"
    try:
        workbook = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"Ошибка при загрузке книги: {e}")
        return

    sheet = workbook.active

    # Определяем стиль NamedStyle
    default_style = NamedStyle(name="default_style")
    default_style.font = Font(name='Calibri', size=11)

    # Применяем стиль default_style ко всему листу
    for row in sheet.iter_rows():
        for cell in row:
            cell.style = default_style

    selected_projects = ["Направление поддержки", "Global", "Global VIP", "Арбитраж", "b2b",
                         "Loyalty Team", "VIP", "Fresh (КЦ)", "ПВЗ", "Соц. сети"]

    training_event_types = ["Тренинг_РГ", "Тренинг_РГ_групповой", "Тренинг_Разбор Ошибок", "Тренинг_РГ_индивидуальный",
                            "Тренинг_Замена РГ", "Тренинг_Наставничество", "Тренинг_Помощь в полях", "Тренинг_РГ (ИПР)",
                            "Тренинг_помощь в Начальном обучении", "Тренинг с наставником"]

    excluded_training_event = "Тренинг_Обучение_НО"
    brand_event_identifier = "Brand Analytics"
    brand_event_identifier2 = "Тренинг_Обработка Отзовик"
    problems_zone_cs = "Тренинг_Проблемные зоны"
    ot_akadem = "Тренинг от отдела обучения"

    training_hours_map = {}
    courses_hours_map = {}
    problem_zone_map = {}
    brand_analytics_hours_map = {}
    ot_akadem_map = {}

    for row in sheet.iter_rows(min_row=2):
        project_name = row[6].value
        event_type = row[14].value
        cell_hours = row[13]

        if event_type == excluded_training_event:
            continue

        if event_type.startswith("Тренинг") and event_type != excluded_training_event:
            if cell_hours.data_type == 's':
                time_value = cell_hours.value
                time_parts = time_value.split(":")
                hours_value = int(time_parts[0])
                minutes_value = int(time_parts[1])
                hours = timedelta(hours=hours_value, minutes=minutes_value)
            else:
                print(f"Не числовое значение в строке: {row[0].row}, значение ячейки: {cell_hours.value}")
                continue

            if is_of_type(event_type, training_event_types):
                if project_name in training_hours_map:
                    training_hours_map[project_name] += hours
                else:
                    training_hours_map[project_name] = hours
            elif brand_event_identifier in event_type or brand_event_identifier2 in event_type:
                if project_name in brand_analytics_hours_map:
                    brand_analytics_hours_map[project_name] += hours
                else:
                    brand_analytics_hours_map[project_name] = hours
            elif ot_akadem in event_type:
                if project_name in ot_akadem_map:
                    ot_akadem_map[project_name] += hours
                else:
                    ot_akadem_map[project_name] = hours
            elif problems_zone_cs in event_type:
                if project_name in problem_zone_map:
                    problem_zone_map[project_name] += hours
                else:
                    problem_zone_map[project_name] = hours
            else:
                if project_name in courses_hours_map:
                    courses_hours_map[project_name] += hours
                else:
                    courses_hours_map[project_name] = hours

    print("+-----------------------+-----------+-----------+-------------+-------------+--------+---------+")
    print("|        Проект         | Тренинги  |   Курсы   | От обучения | Проблем.зоны|   BA   |  Всего  |")
    print("+-----------------------+-----------+-----------+-------------+-------------+--------+---------+")

    for project in selected_projects:
        all_training_hours = training_hours_map.get(project, timedelta())
        courses_hours = courses_hours_map.get(project, timedelta())
        problem_zone_hours = problem_zone_map.get(project, timedelta())
        ot_akadem_hours = ot_akadem_map.get(project, timedelta())
        brand_analytics_hours = brand_analytics_hours_map.get(project, timedelta())
        total_hours = all_training_hours + courses_hours + brand_analytics_hours + problem_zone_hours + ot_akadem_hours

        brand_analytics_duration = format_duration(brand_analytics_hours)
        problem_zone_duration = format_duration(problem_zone_hours)
        ot_akadem_duration = format_duration(ot_akadem_hours)

        if problem_zone_duration == "00:00" and brand_analytics_duration == "00:00" and ot_akadem_duration == "00:00":
            print(f"| {project:<21} | {format_duration(all_training_hours):<9} | {format_duration(courses_hours):<9} | {'':<11} | {'':<11} | {'':<6} | {format_duration(total_hours):<7} |")
        elif problem_zone_duration == "00:00" and brand_analytics_duration == "00:00":
            print(f"| {project:<21} | {format_duration(all_training_hours):<9} | {format_duration(courses_hours):<9} | {ot_akadem_duration:<11} | {'':<11} | {'':<6} | {format_duration(total_hours):<7} |")
        elif ot_akadem_duration == "00:00" and problem_zone_duration == "00:00":
            print(f"| {project:<21} | {format_duration(all_training_hours):<9} | {format_duration(courses_hours):<9} | {'':<11} | {'':<11} | {brand_analytics_duration:<6} | {format_duration(total_hours):<7} |")
        elif ot_akadem_duration == "00:00" and brand_analytics_duration == "00:00":
            print(f"| {project:<21} | {format_duration(all_training_hours):<9} | {format_duration(courses_hours):<9} | {'':<11} | {problem_zone_duration:<11} | {'':<6} | {format_duration(total_hours):<7} |")
        elif problem_zone_duration == "00:00":
            print(f"| {project:<21} | {format_duration(all_training_hours):<9} | {format_duration(courses_hours):<9} | {ot_akadem_duration:<11} | {'':<11} | {brand_analytics_duration:<6} | {format_duration(total_hours):<7} |")
        elif brand_analytics_duration == "00:00":
            print(f"| {project:<21} | {format_duration(all_training_hours):<9} | {format_duration(courses_hours):<9} | {ot_akadem_duration:<11} | {problem_zone_duration:<11} | {'':<6} | {format_duration(total_hours):<7} |")
        elif ot_akadem_duration == "00:00":
            print(f"| {project:<21} | {format_duration(all_training_hours):<9} | {format_duration(courses_hours):<9} | {'':<11} | {problem_zone_duration:<11} | {brand_analytics_duration:<6} | {format_duration(total_hours):<7} |")
        else:
            print(f"| {project:<21} | {format_duration(all_training_hours):<9} | {format_duration(courses_hours):<9} | {ot_akadem_duration:<11} | {problem_zone_duration:<11} | {brand_analytics_duration:<6} | {format_duration(total_hours):<7} |")

    print("+-----------------------+-----------+-----------+-------------+-------------+--------+---------+")

    try:
        workbook.save(filename=file_path)
        print("Файл успешно сохранен.")
    except PermissionError as e:
        print(f"Ошибка доступа: {e}")
    except Exception as e:
        print(f"Ошибка при сохранении книги: {e}")

    input("\nНажмите Enter для завершения программы.")


if __name__ == "__main__":
    main()
